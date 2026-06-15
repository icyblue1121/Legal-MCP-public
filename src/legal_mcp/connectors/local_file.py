"""Local file read-through connector (v0.5.5).

The first source with no external service: it reads a *local* structured or
semi-structured table — CSV, XLSX, JSON / JSONL, or a directory of Markdown files
with YAML frontmatter — and answers the gateway's constrained queries against it.
It exists so the v0.5 onboarding framework has a dependency-free source to drive
end-to-end.

Design choices:

* **Operator parity via in-memory SQLite.** Rather than reimplement the operator
  set in Python, ``query`` loads the (already filtered-to-declared-columns) rows
  into an in-memory SQLite table and reuses :mod:`legal_mcp.connectors.sqlite_filter`
  — so ``eq`` / ``contains`` / ``in`` / ``is_empty`` / ``date_*`` / the virtual
  ``identity`` ``or_fields`` group all behave exactly as on the demo source.
* **Declared catalog, like Feishu.** A file's columns are discoverable
  (``describe_schema``) but only *declared* columns are queryable — a reviewed,
  git-committable security boundary. Undeclared columns never load.
* **Body is out of scope.** Markdown frontmatter keys become columns; the document
  *body* is never read (that is the v0.6 RAG question, explicitly out of scope).
* **Zero-dep core.** CSV and JSON/JSONL use the standard library. XLSX (openpyxl)
  and Markdown frontmatter (PyYAML) are parsed via lazy imports, so the core keeps
  ``dependencies = []`` and those formats simply require the ``local-file`` extra.
"""

from __future__ import annotations

import csv
import json
import sqlite3
from collections.abc import Iterable
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from legal_mcp.connectors.base import (
    ConnectorDomain,
    ConnectorField,
    ConnectorQuery,
    RecordScope,
    SourceTable,
    record_scope_from_dict,
)
from legal_mcp.connectors.sqlite_filter import build_where

SUPPORTED_FORMATS = frozenset({"csv", "xlsx", "json", "jsonl", "md"})


@dataclass(frozen=True)
class LocalFileFieldConfig:
    """A queryable column declared by the operator."""

    name: str
    is_identity: bool = False
    aliases: tuple[str, ...] = ()


@dataclass(frozen=True)
class LocalFileDomainConfig:
    """One gateway domain backed by one local file (or Markdown directory)."""

    name: str
    path: str
    format: str
    fields: tuple[LocalFileFieldConfig, ...]
    relationship_filter_fields: tuple[str, ...] = ()
    record_scope: RecordScope = RecordScope()


@dataclass(frozen=True)
class LocalFileConfig:
    """Declared mapping from gateway domains to local files."""

    domains: tuple[LocalFileDomainConfig, ...]

    @classmethod
    def from_dict(cls, data: dict[str, Any]) -> "LocalFileConfig":
        domains: list[LocalFileDomainConfig] = []
        for raw_domain in data.get("domains") or []:
            raw_domain = raw_domain or {}
            name = raw_domain.get("name")
            path = raw_domain.get("path")
            fmt = (raw_domain.get("format") or "").lower()
            if not name or not path:
                raise ValueError("each local_file domain requires 'name' and 'path'")
            if fmt not in SUPPORTED_FORMATS:
                raise ValueError(
                    f"local_file domain {name!r} has unsupported format {fmt!r}; "
                    f"expected one of {sorted(SUPPORTED_FORMATS)}"
                )
            fields = tuple(
                LocalFileFieldConfig(
                    name=field["name"],
                    is_identity=bool(field.get("is_identity", False)),
                    aliases=tuple(field.get("aliases") or ()),
                )
                for field in (raw_domain.get("fields") or [])
                if field and field.get("name")
            )
            domains.append(
                LocalFileDomainConfig(
                    name=name,
                    path=path,
                    format=fmt,
                    fields=fields,
                    relationship_filter_fields=tuple(
                        raw_domain.get("relationship_filter_fields") or ()
                    ),
                    record_scope=record_scope_from_dict(raw_domain.get("record_scope")),
                )
            )
        return cls(domains=tuple(domains))


class LocalFileConnector:
    """Read-through connector over local files, via an in-memory SQLite stage."""

    name = "local_file"

    def __init__(self, config: LocalFileConfig) -> None:
        self._config = config

    def catalog(self) -> tuple[ConnectorDomain, ...]:
        domains: list[ConnectorDomain] = []
        for domain in self._config.domains:
            fields = tuple(
                ConnectorField(
                    domain=domain.name,
                    name=field.name,
                    is_identity=field.is_identity,
                    aliases=field.aliases,
                )
                for field in domain.fields
            )
            domains.append(
                ConnectorDomain(
                    name=domain.name,
                    table=domain.path,
                    fields=fields,
                    relationship_filter_fields=domain.relationship_filter_fields,
                    record_scope=domain.record_scope,
                )
            )
        return tuple(domains)

    def describe_schema(self) -> tuple[SourceTable, ...]:
        """List each file's *real* columns (values-free) for config scaffolding."""
        tables: list[SourceTable] = []
        for domain in self._config.domains:
            columns = _discover_columns(domain.path, domain.format)
            tables.append(SourceTable(domain=domain.name, table=domain.path, fields=columns))
        return tuple(tables)

    def query(self, query: ConnectorQuery) -> list[dict[str, Any]]:
        domain = self._domain(query.domain)
        known_fields = {field.name for field in domain.fields}
        select_fields = [name for name in query.fields if name in known_fields]
        if not select_fields:
            raise ValueError("query has no known return fields")

        # Validate filter fields up front (declared columns + relationships), so an
        # unknown field is a clear error, not a silent empty result.
        filter_fields = known_fields | set(domain.relationship_filter_fields)
        for query_filter in query.filters:
            for name in (query_filter.or_fields or (query_filter.field,)):
                if name not in filter_fields:
                    raise ValueError(f"unknown filter field: {name}")

        rows = _load_rows(domain.path, domain.format, known_fields)
        return self._query_in_memory(rows, known_fields, select_fields, query)

    def _query_in_memory(
        self,
        rows: list[dict[str, Any]],
        known_fields: set[str],
        select_fields: list[str],
        query: ConnectorQuery,
    ) -> list[dict[str, Any]]:
        # Map each declared field to a safe synthetic column (c0, c1, …) so a file
        # column with spaces/unicode never reaches SQL as an identifier.
        ordered = sorted(known_fields)
        col_of = {name: f"c{index}" for index, name in enumerate(ordered)}

        conn = sqlite3.connect(":memory:")
        conn.row_factory = sqlite3.Row
        try:
            columns_ddl = ", ".join(f"{col_of[name]} text" for name in ordered)
            conn.execute(f"create table data ({columns_ddl})")
            placeholders = ", ".join("?" for _ in ordered)
            conn.executemany(
                f"insert into data ({', '.join(col_of[name] for name in ordered)}) "
                f"values ({placeholders})",
                [tuple(_cell(row.get(name)) for name in ordered) for row in rows],
            )

            params: list[Any] = []

            def column_for(field: str) -> str:
                # A relationship field has no column in a flat file; treat it as a
                # missing column so it simply matches nothing rather than erroring.
                return col_of.get(field, "''")

            where = build_where(query.filters, column_for, params)
            select_list = ", ".join(f"{col_of[name]} as {col_of[name]}" for name in select_fields)
            sql = f"select {select_list} from data"
            if where:
                sql += " where " + " and ".join(where)
            sql += " limit ?"
            params.append(int(query.limit))
            result_rows = conn.execute(sql, params).fetchall()
        finally:
            conn.close()

        # Remap synthetic columns back to the declared field names.
        reverse = {col_of[name]: name for name in select_fields}
        return [
            {reverse[col]: row[col] for col in reverse if row[col] is not None}
            for row in result_rows
        ]

    def _domain(self, name: str) -> LocalFileDomainConfig:
        for domain in self._config.domains:
            if domain.name == name:
                return domain
        raise ValueError(f"unknown domain: {name}")


def _cell(value: Any) -> Any:
    """Normalize a loaded cell to a SQLite-storable scalar (text)."""
    if value is None:
        return None
    if isinstance(value, bool):
        # YAML/JSON booleans -> a stable lowercase string, not SQLite's 1/0.
        return str(value).lower()
    if isinstance(value, str | int | float):
        return value
    if isinstance(value, list | dict):
        # Nested JSON values: store a stable string form so a filter still has
        # something to match; the structure/body itself is out of scope.
        return json.dumps(value, ensure_ascii=False, sort_keys=True)
    # Dates/datetimes (YAML auto-parses ``2024-03-01``) and any other scalar ->
    # their ISO-ish string form, so date_* comparisons work against text.
    return str(value)


# --- format loaders ----------------------------------------------------------


def _discover_columns(path: str, fmt: str) -> tuple[str, ...]:
    if fmt == "csv":
        return _csv_columns(path)
    if fmt == "xlsx":
        return _xlsx_columns(path)
    if fmt in ("json", "jsonl"):
        return _json_columns(path, fmt)
    if fmt == "md":
        return _md_columns(path)
    raise ValueError(f"unsupported format: {fmt}")


def _load_rows(path: str, fmt: str, known_fields: set[str]) -> list[dict[str, Any]]:
    if fmt == "csv":
        raw = _csv_rows(path)
    elif fmt == "xlsx":
        raw = _xlsx_rows(path)
    elif fmt == "json":
        raw = _json_rows(path)
    elif fmt == "jsonl":
        raw = _jsonl_rows(path)
    elif fmt == "md":
        raw = _md_rows(path)
    else:
        raise ValueError(f"unsupported format: {fmt}")
    # Keep only declared columns — undeclared file columns never enter the gateway.
    return [{name: row.get(name) for name in known_fields} for row in raw]


def _csv_columns(path: str) -> tuple[str, ...]:
    with open(path, newline="", encoding="utf-8") as handle:
        reader = csv.reader(handle)
        header = next(reader, [])
    return tuple(str(name) for name in header)


def _csv_rows(path: str) -> list[dict[str, Any]]:
    with open(path, newline="", encoding="utf-8") as handle:
        return [dict(row) for row in csv.DictReader(handle)]


def _json_columns(path: str, fmt: str) -> tuple[str, ...]:
    rows = _json_rows(path) if fmt == "json" else _jsonl_rows(path)
    return _columns_from_records(rows)


def _json_rows(path: str) -> list[dict[str, Any]]:
    data = json.loads(Path(path).read_text(encoding="utf-8"))
    if isinstance(data, dict):
        # Allow a single object or a {"items": [...]} envelope.
        items = data.get("items")
        data = items if isinstance(items, list) else [data]
    if not isinstance(data, list):
        raise ValueError("json source must be an array of objects (or an object)")
    return [record for record in data if isinstance(record, dict)]


def _jsonl_rows(path: str) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for line in Path(path).read_text(encoding="utf-8").splitlines():
        line = line.strip()
        if not line:
            continue
        record = json.loads(line)
        if isinstance(record, dict):
            rows.append(record)
    return rows


def _columns_from_records(rows: Iterable[dict[str, Any]]) -> tuple[str, ...]:
    seen: list[str] = []
    for row in rows:
        for key in row:
            if key not in seen:
                seen.append(str(key))
    return tuple(seen)


def _xlsx_columns(path: str) -> tuple[str, ...]:
    rows = _xlsx_rows(path, header_only=True)
    return tuple(rows[0].keys()) if rows else ()


def _xlsx_rows(path: str, *, header_only: bool = False) -> list[dict[str, Any]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover - environment dependent
        raise ImportError(
            "reading an XLSX local_file source requires openpyxl "
            "(pip install 'legal-mcp[local-file]')"
        ) from exc
    workbook = load_workbook(filename=path, read_only=True, data_only=True)
    try:
        sheet = workbook.active
        iterator = sheet.iter_rows(values_only=True)
        header = next(iterator, None)
        if not header:
            return []
        columns = [str(value) if value is not None else "" for value in header]
        if header_only:
            return [{column: None for column in columns}]
        rows: list[dict[str, Any]] = []
        for values in iterator:
            rows.append(
                {columns[i]: values[i] for i in range(len(columns)) if i < len(values)}
            )
        return rows
    finally:
        workbook.close()


def _md_columns(path: str) -> tuple[str, ...]:
    return _columns_from_records(_md_rows(path))


def _md_rows(path: str) -> list[dict[str, Any]]:
    """One row per ``.md`` file in a directory, from its YAML frontmatter.

    The document body is intentionally ignored (v0.6 scope). ``path`` may be a
    directory of Markdown files or a single Markdown file.
    """
    try:
        import yaml
    except ImportError as exc:  # pragma: no cover - environment dependent
        raise ImportError(
            "reading a Markdown local_file source requires PyYAML "
            "(pip install 'legal-mcp[local-file]')"
        ) from exc
    target = Path(path)
    files = sorted(target.glob("*.md")) if target.is_dir() else [target]
    rows: list[dict[str, Any]] = []
    for file in files:
        frontmatter = _parse_frontmatter(file.read_text(encoding="utf-8"), yaml)
        if frontmatter is not None:
            rows.append(frontmatter)
    return rows


def _parse_frontmatter(text: str, yaml_module: Any) -> dict[str, Any] | None:
    """The leading ``---`` … ``---`` YAML block as a dict, or None if absent."""
    lines = text.splitlines()
    if not lines or lines[0].strip() != "---":
        return None
    closing = next((i for i in range(1, len(lines)) if lines[i].strip() == "---"), None)
    if closing is None:
        return None
    block = "\n".join(lines[1:closing])
    parsed = yaml_module.safe_load(block)
    return parsed if isinstance(parsed, dict) else {}
